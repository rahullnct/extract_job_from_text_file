from django.shortcuts import render

# Create your views here.
import hashlib
import re
import calendar
import re
from .mylib import infer_country_from_location

import json
from html import unescape
from datetime import timedelta
from bs4 import BeautifulSoup

from django.utils import timezone

from datetime import datetime
from io import BytesIO
from pathlib import Path

from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone
from django.utils.text import slugify

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .forms import JobTextUploadForm


# ============================================================
# FINAL JOB EXCEL COLUMNS
# ============================================================

JOB_COLUMNS = [
    "job_id",
    "internal_job_id",
    "job_title",
    "company_name",
    "company_website",
    "job_description",
    "skills_required",
    "experience_required",
    "experience_min_years",
    "experience_max_years",
    "salary_min",
    "salary_max",
    "salary_currency",
    "location",
    "city",
    "state",
    "country",
    "remote_type",
    "employment_type",
    "job_category",
    "posted_date",
    "expiry_date",
    "application_deadline",
    "apply_url",
    "updated_at",
    "fetched_at",
    "is_fake",
    "is_active",
    "source_url",
    "source_type",
    "source_platform",
    "score",
]


# ============================================================
# TECHNOLOGIES / SKILLS
# ============================================================

TECHNOLOGY_PATTERNS = [
    (r"\bJava\b", "Java"),
    (r"\bSpring Boot\b", "Spring Boot"),
    (r"\bSpring\b", "Spring"),
    (r"\bMaven\b", "Maven"),

    (r"\bReact(?:JS|\.js)?\b", "React"),
    (r"\bNext\s*\.?\s*JS\b", "Next.js"),
    (r"\bTypeScript\b", "TypeScript"),
    (r"\bJavaScript\b", "JavaScript"),

    (r"\bAngular\b", "Angular"),
    (r"\bVue(?:JS|\.js)?\b", "Vue.js"),

    (r"\bNode(?:JS|\.js)?\b", "Node.js"),
    (r"\bExpress(?:JS|\.js)?\b", "Express.js"),

    (r"\bPython\b", "Python"),
    (r"\bDjango\b", "Django"),
    (r"\bFlask\b", "Flask"),

    (r"\bTerraform\b", "Terraform"),
    (r"\bKubernetes\b", "Kubernetes"),
    (r"\bDocker\b", "Docker"),

    (r"\bPostgreSQL\b", "PostgreSQL"),
    (r"\bSupabase\b", "Supabase"),
    (r"\bMongoDB\b", "MongoDB"),
    (r"\bMySQL\b", "MySQL"),
    (r"\bOracle\b", "Oracle"),
    (r"\bSQL\b", "SQL"),

    (r"\bAzure\b", "Azure"),
    (r"\bAWS\b", "AWS"),
    (r"\bGCP\b", "GCP"),
    (r"\bGoogle Cloud\b", "Google Cloud"),

    (r"\bCI\s*/\s*CD\b", "CI/CD"),
    (r"\bJira\b", "Jira"),
    (r"\bGit\b", "Git"),
    (r"\bDevOps\b", "DevOps"),

    (r"\bREST(?:ful)?\s*API(?:s)?\b", "REST API"),
    (r"\bLLM(?:s)?\b", "LLM"),
    (r"\bNLP\b", "NLP"),
    (r"\bPyTorch\b", "PyTorch"),
    (r"\bTensorFlow\b", "TensorFlow"),
    (r"\bLangChain\b", "LangChain"),

    (r"\bFastAPI\b", "FastAPI"),

    (r"\bOpenAI\b", "OpenAI"),
    (r"\bHugging Face\b", "Hugging Face"),

    (r"\bRAG\b", "RAG"),

    (r"\bPinecone\b", "Pinecone"),
    (r"\bFAISS\b", "FAISS"),
    (r"\bWeaviate\b", "Weaviate"),
    (r"\bChromaDB\b", "ChromaDB"),

    (r"\bAirflow\b", "Airflow"),
    (r"\bPrefect\b", "Prefect"),

    (r"\bNoSQL\b", "NoSQL"),

    (r"\bMLOps\b", "MLOps"),

    (r"\bLLaMA\b", "LLaMA"),
    (r"\bFalcon\b", "Falcon"),
    (r"\bMistral\b", "Mistral"),

    (r"\bWordPress\b", "WordPress"),
    (r"\bShopify\b", "Shopify"),
    (r"\bElementor\b", "Elementor"),
    (r"\bBricks\b", "Bricks"),
    (r"\bUI\s*/\s*UX\b", "UI/UX"),
    (r"\bWeb Development\b", "Web Development"),
    (r"\bE-commerce\b", "E-commerce"),
]

POSTED_TIME_PATTERN = re.compile(
    r"^"
    r"(\d+)"
    r"\s+"
    r"(minute|minutes|hour|hours|day|days|week|weeks|month|months|year|years)"
    r"\s+ago"
    r"$",
    re.IGNORECASE,
)



def subtract_months(date_value, months):
    """
    Subtract calendar months safely.

    Example:
    2026-08-31 minus 1 month -> 2026-07-31
    """

    total_months = (
        date_value.year * 12
        + date_value.month
        - 1
        - months
    )

    year = total_months // 12
    month = total_months % 12 + 1

    last_day = calendar.monthrange(
        year,
        month,
    )[1]

    day = min(
        date_value.day,
        last_day,
    )

    return date_value.replace(
        year=year,
        month=month,
        day=day,
    )

def convert_relative_posted_date(value):
    """
    Convert:

    2 months ago
    3 weeks ago
    1 day ago

    into YYYY-MM-DD.

    The calculation uses the date when the TXT is processed.
    """

    if not value:
        return ""

    match = POSTED_TIME_PATTERN.fullmatch(
        value.strip()
    )

    if not match:
        return ""

    amount = int(
        match.group(1)
    )

    unit = (
        match.group(2)
        .lower()
    )

    now = timezone.localtime(
        timezone.now()
    )

    if unit in {
        "minute",
        "minutes",
    }:
        result = now - timedelta(
            minutes=amount
        )

    elif unit in {
        "hour",
        "hours",
    }:
        result = now - timedelta(
            hours=amount
        )

    elif unit in {
        "day",
        "days",
    }:
        result = now - timedelta(
            days=amount
        )

    elif unit in {
        "week",
        "weeks",
    }:
        result = now - timedelta(
            weeks=amount
        )

    elif unit in {
        "month",
        "months",
    }:
        result = subtract_months(
            now,
            amount,
        )

    elif unit in {
        "year",
        "years",
    }:
        result = subtract_months(
            now,
            amount * 12,
        )

    else:
        return ""

    return result.strftime(
        "%Y-%m-%d"
    )


# ============================================================
# FILE DECODING
# ============================================================

def decode_uploaded_file(uploaded_file):

    raw_data = uploaded_file.read()

    for encoding in [
        "utf-8-sig",
        "utf-8",
        "cp1252",
        "latin-1",
    ]:

        try:
            return raw_data.decode(encoding)

        except UnicodeDecodeError:
            pass

    return raw_data.decode(
        "utf-8",
        errors="ignore",
    )


# ============================================================
# NORMALIZE TEXT
# ============================================================

def normalize_text(text):

    text = text.replace("\r\n", "\n")
    text = text.replace("\r", "\n")
    text = text.replace("\xa0", " ")

    lines = []

    for line in text.split("\n"):

        line = line.strip()

        lines.append(line)

    return "\n".join(lines)


# ============================================================
# REMOVE SIMILAR JOBS
# ============================================================



# ============================================================
# GENERIC LABEL VALUE
# ============================================================

def get_label_value(text, labels):

    if isinstance(labels, str):
        labels = [labels]

    for label in labels:

        pattern = (
            rf"(?im)^\s*"
            rf"{re.escape(label)}"
            rf"\s*:?\s*"
            rf"(?:\n\s*)?"
            rf"([^\n]+)"
        )

        match = re.search(
            pattern,
            text,
        )

        if match:

            value = (
                match.group(1)
                .strip()
            )

            if value:
                return value

    return ""


# ============================================================
# SECTION EXTRACTOR
# ============================================================

def extract_section(
    text,
    start_heading,
    end_headings,
):

    lower_text = text.lower()

    start = lower_text.find(
        start_heading.lower()
    )

    if start == -1:
        return ""

    start += len(start_heading)

    section = text[start:]

    lower_section = section.lower()

    end_positions = []

    for heading in end_headings:

        position = lower_section.find(
            heading.lower()
        )

        if position != -1:
            end_positions.append(
                position
            )

    if end_positions:

        section = section[
            :min(end_positions)
        ]

    return section.strip()




# ============================================================
# COMPANY NAME
# ============================================================



# ============================================================
# JOB DESCRIPTION
# ============================================================


def extract_job_description(text):

    description = extract_section(
        text,
        "Job Description",
        [
            "Other Details",
            "Recruiter Details",
            "About Recruiter",
        ],
    )

    description = re.sub(
        r"\n{3,}",
        "\n\n",
        description,
    )

    return description.strip()



# ============================================================
# EXPERIENCE
# ============================================================


def extract_experience(text):

    patterns = [
        r"(?:Experience\s*:\s*)?"
        r"(\d+(?:\.\d+)?)"
        r"\s*to\s*"
        r"(\d+(?:\.\d+)?)"
        r"\s*(?:Yrs?|Years?)",

        r"(?:Experience\s*:\s*)?"
        r"(\d+(?:\.\d+)?)"
        r"\s*-\s*"
        r"(\d+(?:\.\d+)?)"
        r"\s*(?:Yrs?|Years?)",
    ]

    for pattern in patterns:

        match = re.search(
            pattern,
            text,
            flags=re.IGNORECASE,
        )

        if match:

            minimum = float(
                match.group(1)
            )

            maximum = float(
                match.group(2)
            )

            if minimum.is_integer():
                minimum = int(minimum)

            if maximum.is_integer():
                maximum = int(maximum)

            return (
                f"{minimum} to {maximum} Years",
                minimum,
                maximum,
            )

    # Example:
    # 5 Yrs

    match = re.search(
        r"\b(\d+(?:\.\d+)?)\s*Yrs?\b",
        text,
        flags=re.IGNORECASE,
    )

    if match:

        years = float(
            match.group(1)
        )

        if years.is_integer():
            years = int(years)

        return (
            f"{years} Yrs",
            years,
            years,
        )

    return "", "", ""

def extract_linkedin_employment_type(text):

    allowed_types = {
        "full-time": "Full-time",
        "full time": "Full-time",

        "part-time": "Part-time",
        "part time": "Part-time",

        "internship": "Internship",

        "contract": "Contract",

        "temporary": "Temporary",

        "freelance": "Freelance",
    }

    lines = [
        line.strip()
        for line in text.split("\n")
        if line.strip()
    ]

    for line in lines[:30]:

        normalized = line.lower()

        if normalized in allowed_types:

            return allowed_types[
                normalized
            ]

    # Fallback from description:
    #
    # Job Type: Full Time

    value = get_label_value(
        text,
        "Job Type",
    )

    if value:

        lower_value = value.lower()

        return allowed_types.get(
            lower_value,
            value,
        )

    return ""

def parse_linkedin_location(location):

    if not location:
        return "", "", ""

    location = re.sub(
        r"\([^)]*\)",
        "",
        location,
    ).strip()

    parts = [
        part.strip()
        for part in location.split(",")
        if part.strip()
    ]

    city = ""
    state = ""
    country = ""

    if len(parts) >= 1:
        city = parts[0]

    if len(parts) >= 2:
        state = parts[1]

    if len(parts) >= 3:
        country = parts[-1]

    return (
        city,
        state,
        country,
    )

def detect_linkedin_remote_type(text):

    lower = text.lower()

    if (
        "(remote)" in lower
        or " remote " in f" {lower} "
    ):
        return "Remote"

    if (
        "(hybrid)" in lower
        or " hybrid " in f" {lower} "
    ):
        return "Hybrid"

    if (
        "(on-site)" in lower
        or "(onsite)" in lower
        or "on-site" in lower
    ):
        return "Onsite"

    return ""

def extract_linkedin_job_category(
    text,
    job_title,
    job_description,
):

    explicit_category = (
        get_label_value(
            text,
            "Job Category",
        )
    )

    if explicit_category:
        return explicit_category

    return detect_job_category(
        job_title,
        job_description,
    )

# ============================================================
# SALARY
# ============================================================


def extract_salary(text):

    # Example:
    # Rs 3.0 - 6 Lakh/Yr
    #
    # Convert to actual annual INR:
    # 3 Lakh = 300000

    pattern = (
        r"(?:Rs\.?|₹)?\s*"
        r"(\d+(?:\.\d+)?)"
        r"\s*(?:-|to)\s*"
        r"(\d+(?:\.\d+)?)"
        r"\s*(?:Lakh|LPA)"
    )

    match = re.search(
        pattern,
        text,
        flags=re.IGNORECASE,
    )

    if match:

        minimum = (
            float(match.group(1))
            * 100000
        )

        maximum = (
            float(match.group(2))
            * 100000
        )

        return (
            int(minimum),
            int(maximum),
            "INR",
        )

    return "", "", ""


# ============================================================
# LOCATION
# ============================================================



# ============================================================
# REMOTE TYPE
# ============================================================


def detect_remote_type(text):

    lower = text.lower()

    if (
        "wfh/wfa" in lower
        or "work from home" in lower
        or "work from anywhere" in lower
        or re.search(
            r"\bremote\b",
            lower,
        )
    ):
        return "Remote"

    if "hybrid" in lower:
        return "Hybrid"

    if (
        "work from office" in lower
        or "onsite" in lower
        or "on-site" in lower
    ):
        return "Onsite"

    return ""


# ============================================================
# EMPLOYMENT TYPE
# ============================================================


def extract_employment_type(text):

    value = get_label_value(
        text,
        "Job Type",
    )

    lower = value.lower()

    mapping = [
        ("full time", "Full-time"),
        ("full-time", "Full-time"),

        ("part time", "Part-time"),
        ("part-time", "Part-time"),

        ("internship", "Internship"),

        ("contract", "Contract"),

        ("temporary", "Temporary"),

        ("freelance", "Freelance"),
    ]

    for keyword, normalized in mapping:

        if keyword in lower:
            return normalized

    return value


# ============================================================
# SKILLS
# ============================================================

def extract_skills(job_description):

    skills = []

    for pattern, skill_name in TECHNOLOGY_PATTERNS:

        if re.search(
            pattern,
            job_description,
            flags=re.IGNORECASE,
        ):

            if skill_name not in skills:
                skills.append(
                    skill_name
                )

    return "; ".join(skills)


# ============================================================
# JOB CATEGORY
# ============================================================

def detect_job_category(
    title,
    description,
):

    text = (
        f"{title} {description}"
    ).lower()

    it_keywords = [
        "developer",
        "software",
        "java",
        "python",
        "engineer",
        "frontend",
        "backend",
        "full stack",
        "devops",
        "cloud",
        "data scientist",
        "cybersecurity",
        "database",
    ]

    finance_keywords = [
        "finance",
        "accountant",
        "accounting",
        "financial analyst",
        "banking",
    ]

    marketing_keywords = [
        "marketing",
        "seo",
        "digital marketing",
        "brand manager",
    ]

    hr_keywords = [
        "human resource",
        "human resources",
        "hr manager",
        "recruiter",
        "talent acquisition",
    ]

    sales_keywords = [
        "sales",
        "business development",
        "lead generation",
    ]

    for keyword in it_keywords:

        if keyword in text:
            return "IT"

    for keyword in finance_keywords:

        if keyword in text:
            return "Finance"

    for keyword in marketing_keywords:

        if keyword in text:
            return "Marketing"

    for keyword in hr_keywords:

        if keyword in text:
            return "Human Resources"

    for keyword in sales_keywords:

        if keyword in text:
            return "Sales"

    return "Other"


# ============================================================
# POSTED DATE
# ============================================================

def extract_posted_date(text):

    match = re.search(
        r"Date:\s*(\d{1,2}/\d{1,2}/\d{4})",
        text,
        flags=re.IGNORECASE,
    )

    if not match:
        return ""

    raw_date = match.group(1)

    try:

        parsed = datetime.strptime(
            raw_date,
            "%m/%d/%Y",
        )

        return parsed.strftime(
            "%Y-%m-%d"
        )

    except ValueError:

        return raw_date


# ============================================================
# WEBSITE
# ============================================================

def extract_company_website(text):

    return get_label_value(
        text,
        [
            "Company Website",
            "Official Website",
            "Website",
        ],
    )


# ============================================================
# SOURCE PLATFORM DETECTION
# ============================================================

def detect_source_platform(
    source_html,
    source_url="",
):
    """
    Detect the actual job portal.

    Priority:

    1. User-provided source URL
    2. Strong portal-specific HTML markers
    3. Multiple-marker scoring

    Do not identify a portal only because its social-media
    URL appears somewhere in another portal's footer.
    """

    source_url_lower = (
        source_url or ""
    ).lower()

    html_lower = (
        source_html or ""
    ).lower()

    # =====================================================
    # 1. SOURCE URL
    # =====================================================

    url_platforms = [
        (
            "Shine",
            [
                "shine.com",
            ],
        ),

        (
            "Indeed",
            [
                "indeed.com",
            ],
        ),
        (
            "Cutshort",
            [
                "cutshort.io/job/",
                "cutshort.io/job",
            ],
        ),
        (
            "LinkedIn",
            [
                "linkedin.com/jobs",
            ],
        ),

        (
            "Naukri",
            [
                "naukri.com/job-listings",
                "naukri.com/job",
            ],
        ),

        (
            "Foundit",
            [
                "foundit.in/job",
                "foundit.in/job-vacancy",
            ],
        ),

        (
            "Glassdoor",
            [
                "glassdoor.co.in/job-listing",
                "glassdoor.com/job-listing",
            ],
        ),

        (
            "Internshala",
            [
                "internshala.com/job",
                "internshala.com/internship",
            ],
        ),

        (
            "TimesJobs",
            [
                "timesjobs.com/job-detail",
            ],
        ),

        (
            "Freshersworld",
            [
                "freshersworld.com/jobs",
            ],
        ),

        (
            "Greenhouse",
            [
                "greenhouse.io",
            ],
        ),

        (
            "Lever",
            [
                "lever.co",
            ],
        ),
    ]

    for platform, patterns in url_platforms:

        for pattern in patterns:

            if pattern in source_url_lower:
                return platform

    # =====================================================
    # 2. STRONG HTML MARKERS
    # =====================================================

    platform_markers = {

        "Indeed": [
            'id="job-full-details"',
            "viewjobssrroot",
            "jobsearch-rightpane",
            "jobsearch-jobinfoheader-title",
            "jobsearch-viewjobcontainerwrapper",
            "data-indeed-apply-jk",
            'id="jobdescriptiontext"',
        ],

        "Shine": [
            "jdleft_jdbodyleft__",
            "jdleft_jdleft__",
            "jdright_jdbodyright__",
            "staticcand.shine.com",
        ],
        "Cutshort": [
            "sc-8f06d440-0",
            "jdluxx",
            "cutshort.io/job/",
            "cdn.cutshort.io",
        ],
        "LinkedIn": [
            "jobs-details__main-content",
            "job-details-jobs-unified-top-card",
            "jobs-description__content",
            "linkedin corporation",
            "jobs-search__job-details",
        ],

        "Naukri": [
            "styles_jd-header",
            "jobdesc",
            "styles_job-desc-container",
            "naukri.com/job-listings",
        ],

        "Foundit": [
            "job-description",
            "foundit.in",
            "jobdetail",
        ],

        "Internshala": [
            "internship_details",
            "internship_heading",
            "internshala",
        ],

        "TimesJobs": [
            "job-detail",
            "jd-jobid",
            "timesjobs",
        ],

        "Freshersworld": [
            "job-container",
            "freshersworld",
        ],

        "Glassdoor": [
            "jobdetails",
            "glassdoor",
        ],
    }

    # =====================================================
    # 3. SCORE EACH PORTAL
    # =====================================================

    scores = {}

    for platform, markers in (
        platform_markers.items()
    ):

        score = 0

        for marker in markers:

            if marker in html_lower:

                score += 1

        scores[platform] = score

    # =====================================================
    # REQUIRE MULTIPLE MARKERS WHEN POSSIBLE
    # =====================================================

    best_platform = ""
    best_score = 0

    for platform, score in scores.items():

        if score > best_score:

            best_platform = platform
            best_score = score

    if best_score >= 2:

        return best_platform

    # Some extremely strong markers can identify
    # the portal by themselves.

    if 'id="job-full-details"' in html_lower:
        return "Indeed"

    if "jdleft_jdbodyleft__" in html_lower:
        return "Shine"
    if (
        "sc-8f06d440-0" in html_lower
        and "cutshort.io" in html_lower
    ):
        return "Cutshort"

    return "Other"

# ============================================================
# GENERIC JOBPOSTING JSON-LD EXTRACTOR
# ============================================================

def extract_jobposting_json_ld(
    source_html,
):
    """
    Find a Schema.org JobPosting object from any website.

    This is NOT Shine-specific.

    It can potentially work with:
    Shine,
    Naukri,
    company career pages,
    ATS pages,
    and other job portals.
    """

    if not source_html:
        return {}

    soup = BeautifulSoup(
        source_html,
        "html.parser",
    )

    scripts = soup.find_all(
        "script",
        attrs={
            "type": "application/ld+json",
        },
    )

    def find_jobposting(
        value,
    ):

        if isinstance(
            value,
            dict,
        ):

            object_type = (
                value.get(
                    "@type"
                )
            )

            if object_type == "JobPosting":

                return value

            if (
                isinstance(
                    object_type,
                    list,
                )
                and "JobPosting"
                in object_type
            ):

                return value

            # JSON-LD often stores objects under:
            #
            # @graph
            #
            # so search recursively.

            for child in value.values():

                result = find_jobposting(
                    child
                )

                if result:
                    return result

        elif isinstance(
            value,
            list,
        ):

            for child in value:

                result = find_jobposting(
                    child
                )

                if result:
                    return result

        return {}

    for script in scripts:

        raw_json = (
            script.string
            or script.get_text()
        )

        if not raw_json:
            continue

        try:

            data = json.loads(
                raw_json
            )

        except (
            json.JSONDecodeError,
            TypeError,
        ):

            continue

        result = find_jobposting(
            data
        )

        if result:
            return result

    return {}

# ============================================================
# SOURCE TYPE
# ============================================================

def isolate_linkedin_job_text(text):
    """
    Keep only the currently opened LinkedIn job.

    Everything before:
        Get job alerts for this search

    belongs to search results/navigation.

    Everything after:
        Job search faster with Premium

    is unrelated LinkedIn UI.
    """

    start_markers = [
        "Get job alerts for this search",
    ]

    end_markers = [
        "Job search faster with Premium",
        "MessagingYou are on the messaging overlay",
    ]

    lower_text = text.lower()

    start_position = None
    matched_marker = ""

    for marker in start_markers:

        position = lower_text.find(
            marker.lower()
        )

        if position != -1:

            start_position = position
            matched_marker = marker
            break

    if start_position is not None:

        text = text[
            start_position
            + len(matched_marker):
        ]

    lower_text = text.lower()

    end_positions = []

    for marker in end_markers:

        position = lower_text.find(
            marker.lower()
        )

        if position != -1:
            end_positions.append(
                position
            )

    if end_positions:

        text = text[
            :min(end_positions)
        ]

    return text.strip()

def extract_linkedin_header(text):
    """
    Extract LinkedIn selected-job header.

    Supported examples:

    Example 1:
        Company logo for, Flourishers Edge.
        Flourishers Edge
        Front End Developer
        Bhopal, Madhya Pradesh, India · 1 month ago ...

    Example 2:
        Company logo for, PwC India.
        PwC India
        Senior Associate
        Bhopal, Madhya Pradesh, India · Reposted 3 days ago ...

    Example 3:
        Netlink Computer Inc
        Python Developer
        Bhopal, Madhya Pradesh, India · 4 days ago ...

    Returns:
        job_title
        company_name
        location
        posted_date_raw
    """

    lines = [
        line.strip()
        for line in text.split("\n")
        if line.strip()
    ]

    if len(lines) < 2:
        return "", "", "", ""

    # =====================================================
    # REMOVE LINKEDIN COMPANY-LOGO ACCESSIBILITY TEXT
    # =====================================================
    #
    # Example:
    #
    # Company logo for, Flourishers Edge.
    # Flourishers Edge
    #
    # The first line is NOT the company name.
    # =====================================================

    cleaned_lines = []

    company_logo_pattern = re.compile(
        r"^Company\s+logo\s+for,\s*(.+?)\.?$",
        re.IGNORECASE,
    )

    for line in lines:

        match = company_logo_pattern.match(
            line
        )

        if match:
            # Skip accessibility/UI logo line.
            continue

        cleaned_lines.append(
            line
        )

    lines = cleaned_lines

    if len(lines) < 2:
        return "", "", "", ""

    # =====================================================
    # COMPANY + JOB TITLE
    # =====================================================
    #
    # After removing logo text LinkedIn becomes:
    #
    # Flourishers Edge
    # Front End Developer
    #
    # OR:
    #
    # Netlink Computer Inc
    # Python Developer
    # =====================================================

    company_name = lines[0]
    job_title = lines[1]

    location = ""
    posted_date_raw = ""

    # =====================================================
    # LOCATION + POSTED DATE
    # =====================================================

    for line in lines[2:15]:

        # Supports:
        #
        # 4 days ago
        # 1 month ago
        # Reposted 3 days ago
        # Posted 2 weeks ago

        relative_match = re.search(
            r"\b"
            r"(?:Reposted\s+|Posted\s+)?"
            r"(\d+)\s+"
            r"(minute|minutes|hour|hours|"
            r"day|days|week|weeks|"
            r"month|months|year|years)"
            r"\s+ago"
            r"\b",
            line,
            flags=re.IGNORECASE,
        )

        if relative_match:

            # We only want:
            #
            # 3 days ago
            #
            # instead of:
            #
            # Reposted 3 days ago

            posted_date_raw = (
                f"{relative_match.group(1)} "
                f"{relative_match.group(2)} ago"
            )

            # LinkedIn location is before first:
            # ·
            if "·" in line:

                location = (
                    line.split(
                        "·",
                        1,
                    )[0]
                    .strip()
                )

            break

    return (
        job_title,
        company_name,
        location,
        posted_date_raw,
    )

def extract_linkedin_job_description(text):
    """
    Extract everything after 'About the job'.
    """

    lower_text = text.lower()

    marker = "about the job"

    position = lower_text.find(
        marker
    )

    if position == -1:
        return ""

    description = text[
        position + len(marker):
    ]

    end_markers = [
        "Job search faster with Premium",
        "See jobs where you’re a top applicant",
        "See jobs where you're a top applicant",
    ]

    lower_description = (
        description.lower()
    )

    end_positions = []

    for end_marker in end_markers:

        end_position = (
            lower_description.find(
                end_marker.lower()
            )
        )

        if end_position != -1:
            end_positions.append(
                end_position
            )

    if end_positions:

        description = description[
            :min(end_positions)
        ]

    description = re.sub(
        r"\n{3,}",
        "\n\n",
        description,
    )

    return description.strip()


def determine_source_type(
    source_platform,
):

    job_portals = {
        "Shine",
        "LinkedIn",
        "Indeed",
        "Naukri",
        "Cutshort",
    }

    ats_sources = {
        "Greenhouse",
        "Lever",
    }

    if source_platform in job_portals:
        return "job_portal"

    if source_platform in ats_sources:
        return "ats_api"

    return "career_page"


# ============================================================
# INITIAL SCORE
# ============================================================

def get_source_score(source_type):

    scores = {
        "ats_api": 60,
        "career_page": 40,
        "job_portal": 30,
    }

    return scores.get(
        source_type,
        0,
    )


# ============================================================
# STABLE JOB IDENTIFIER
# ============================================================

def generate_internal_job_id(
    main_text,
    source_url,
):

    if source_url:
        identity_source = source_url
    else:
        identity_source = main_text

    digest = hashlib.sha256(
        identity_source.encode(
            "utf-8"
        )
    ).hexdigest()

    return digest[:32]


# ============================================================
# VIEW SOURCE UNWRAPPER
# ============================================================

def unwrap_view_source_file(raw_text):
    """
    Convert a Chrome-saved View Source page back into
    the original page HTML.

    Also supports a TXT file that already contains
    normal raw HTML.
    """

    soup = BeautifulSoup(
        raw_text,
        "html.parser",
    )

    source_lines = soup.select(
        "td.line-content"
    )

    # If Chrome View Source wrapper exists,
    # rebuild the original source line-by-line.
    if source_lines:

        original_html = "\n".join(
            line.get_text(
                "",
                strip=False,
            )
            for line in source_lines
        )

        return original_html

    # Otherwise the uploaded TXT already
    # contains normal HTML source.
    return raw_text

# ============================================================
# ISOLATE INDEED SELECTED JOB RIGHT PANE
# ============================================================

def isolate_indeed_job_section(
    source_html,
):
    """
    Extract the complete selected Indeed job area.

    Primary starting container:

    <div class="jobsearch-RightPane ...">

    We intentionally match only the stable class:
        jobsearch-RightPane

    We DO NOT depend on generated classes such as:
        css-6iabie
        eu4oa1w0

    because Indeed may change those values.

    Fallbacks are kept so older Indeed source files
    continue working.
    """

    if not source_html:
        return ""

    soup = BeautifulSoup(
        source_html,
        "html.parser",
    )

    # =====================================================
    # 1. PRIMARY:
    # COMPLETE INDEED RIGHT PANE
    # =====================================================

    right_pane = soup.select_one(
        "div.jobsearch-RightPane"
    )

    if right_pane:

        return str(
            right_pane
        )

    # =====================================================
    # 2. FALLBACK:
    # OLD SELECTED JOB SECTION
    # =====================================================

    job_section = soup.find(
        "section",
        id="job-full-details",
    )

    if job_section:

        return str(
            job_section
        )

    # =====================================================
    # 3. FALLBACK:
    # INDEED SSR ROOT
    # =====================================================

    view_job_root = soup.find(
        id="viewJobSSRRoot"
    )

    if view_job_root:

        parent_right_pane = (
            view_job_root.find_parent(
                "div",
                class_=lambda value:
                    value
                    and (
                        "jobsearch-RightPane"
                        in str(value)
                    )
            )
        )

        if parent_right_pane:

            return str(
                parent_right_pane
            )

        parent_section = (
            view_job_root.find_parent(
                "section"
            )
        )

        if parent_section:

            return str(
                parent_section
            )

        return str(
            view_job_root
        )

    # =====================================================
    # 4. FINAL FALLBACK:
    # JOB DESCRIPTION
    # =====================================================

    description_tag = soup.find(
        id="jobDescriptionText"
    )

    if description_tag:

        parent_right_pane = (
            description_tag.find_parent(
                "div",
                class_=lambda value:
                    value
                    and (
                        "jobsearch-RightPane"
                        in str(value)
                    )
            )
        )

        if parent_right_pane:

            return str(
                parent_right_pane
            )

        return str(
            description_tag.parent
        )

    return ""

# ============================================================
# INDEED SOURCE LOCATION
# ============================================================

def parse_indeed_source_location(
    location,
    country_code="",
):
    """
    Examples:

    Mumbai, Maharashtra
        city    = Mumbai
        state   = Maharashtra

    Borivali, Mumbai, Maharashtra
        city    = Borivali, Mumbai
        state   = Maharashtra

    Borivali, Mumbai, Maharashtra, India
        city    = Borivali, Mumbai
        state   = Maharashtra
        country = India

    Remote
        city    = ""
        state   = ""
    """

    country_map = {
        "IN": "India",
        "US": "United States",
        "GB": "United Kingdom",
        "UK": "United Kingdom",
        "CA": "Canada",
        "AU": "Australia",
        "AE": "United Arab Emirates",
        "SG": "Singapore",
    }

    country = ""

    if country_code:

        country = country_map.get(
            country_code.upper(),
            country_code,
        )

    if not location:

        return "", "", country

    location = location.strip()

    # Pure remote job has no physical city/state.
    if location.lower() == "remote":

        return "", "", country

    # Remove work-mode suffix.
    #
    # Mumbai, Maharashtra•Remote
    # ->
    # Mumbai, Maharashtra

    location = re.sub(
        r"\s*[•·]\s*"
        r"(?:Remote|Hybrid|On-site|Onsite)"
        r"\s*$",
        "",
        location,
        flags=re.IGNORECASE,
    )

    # Remote in Mumbai, Maharashtra
    # ->
    # Mumbai, Maharashtra

    location = re.sub(
        r"^Remote\s+in\s+",
        "",
        location,
        flags=re.IGNORECASE,
    )

    # Hybrid work in Mumbai, Maharashtra
    # ->
    # Mumbai, Maharashtra

    location = re.sub(
        r"^Hybrid\s+work\s+in\s+",
        "",
        location,
        flags=re.IGNORECASE,
    )

    location = location.strip()

    parts = [
        part.strip()
        for part in location.split(",")
        if part.strip()
    ]

    if not parts:
        return "", "", country

    # -----------------------------------------------
    # Country explicitly present in location.
    # -----------------------------------------------

    known_country_names = {
        "india",
        "united states",
        "united kingdom",
        "canada",
        "australia",
        "singapore",
        "united arab emirates",
    }

    if (
        len(parts) >= 3
        and parts[-1].lower()
        in known_country_names
    ):

        country = parts[-1]
        state = parts[-2]

        city = ", ".join(
            parts[:-2]
        )

        return (
            city,
            state,
            country,
        )

    # -----------------------------------------------
    # No explicit country.
    #
    # Last item = state
    # Everything before = city
    # -----------------------------------------------

    if len(parts) == 1:

        return (
            parts[0],
            "",
            country,
        )

    state = parts[-1]

    city = ", ".join(
        parts[:-1]
    )

    return (
        city,
        state,
        country,
    )

# ============================================================
# INDEED SOURCE EMPLOYMENT TYPE
# ============================================================

def extract_indeed_source_employment_type(
    job_section_text,
):
    """
    Extract one or multiple employment types.

    Example:

    Job type

    Permanent
    Full-time

    Location

    -> Permanent, Full-time
    """

    employment_pattern = re.compile(
        r"""
        \b(
            Full[\s-]?time
            |
            Part[\s-]?time
            |
            Internship
            |
            Contract
            |
            Temporary
            |
            Freelance
            |
            Permanent
        )\b
        """,
        flags=(
            re.IGNORECASE
            | re.VERBOSE
        ),
    )

    # Try to isolate Job type section first.
    block_match = re.search(
        r"""
        Job\s+type
        \s*
        (.*?)
        (?=
            Location
            |
            Benefits
            |
            Full\s+job\s+description
            |
            Qualifications
            |
            $
        )
        """,
        job_section_text,
        flags=(
            re.IGNORECASE
            | re.DOTALL
            | re.VERBOSE
        ),
    )

    if block_match:

        search_text = (
            block_match.group(1)
        )

    else:

        search_text = (
            job_section_text
        )

    matches = (
        employment_pattern.findall(
            search_text
        )
    )

    normalization = {
        "full time": "Full-time",
        "part time": "Part-time",
        "internship": "Internship",
        "contract": "Contract",
        "temporary": "Temporary",
        "freelance": "Freelance",
        "permanent": "Permanent",
    }

    found_types = []

    for match in matches:

        normalized_key = (
            match
            .lower()
            .replace(
                "-",
                " ",
            )
        )

        normalized_key = re.sub(
            r"\s+",
            " ",
            normalized_key,
        ).strip()

        employment_type = (
            normalization.get(
                normalized_key,
                "",
            )
        )

        if (
            employment_type
            and employment_type
            not in found_types
        ):

            found_types.append(
                employment_type
            )

    return ", ".join(
        found_types
    )

# ============================================================
# INDEED SOURCE SALARY
# ============================================================

def extract_indeed_source_salary(
    text,
):
    """
    Examples:

    ₹3,00,000 - ₹4,50,000 a year

    ₹30,000 - ₹50,000 a month

    ₹50 - ₹100 an hour

    Monthly salary is converted to annual salary.

    Hourly salary is kept as-is because annual conversion
    cannot be done safely without knowing working hours.
    """

    pattern = re.compile(
        r"""
        ₹\s*
        ([\d,]+(?:\.\d+)?)
        \s*
        (?:-|–|—|to)
        \s*
        ₹?\s*
        ([\d,]+(?:\.\d+)?)
        \s*
        (?:a|per)?
        \s*
        (hour|day|week|month|year|annum)
        """,
        flags=(
            re.IGNORECASE
            | re.VERBOSE
        ),
    )

    match = pattern.search(
        text
    )

    if not match:

        return "", "", ""

    minimum = float(
        match.group(1).replace(
            ",",
            "",
        )
    )

    maximum = float(
        match.group(2).replace(
            ",",
            "",
        )
    )

    period = (
        match.group(3)
        .lower()
    )

    # Monthly -> annual.
    if period == "month":

        minimum *= 12
        maximum *= 12

    if minimum.is_integer():
        minimum = int(minimum)

    if maximum.is_integer():
        maximum = int(maximum)

    return (
        minimum,
        maximum,
        "INR",
    )

# ============================================================
# INDEED SOURCE POSTED DATE
# ============================================================

def extract_indeed_source_posted_date(
    text,
):

    match = re.search(
        r"\b"
        r"(\d+)"
        r"(\+)?"
        r"\s+"
        r"(minute|minutes|hour|hours|"
        r"day|days|week|weeks|"
        r"month|months|year|years)"
        r"\s+ago"
        r"\b",
        text,
        flags=re.IGNORECASE,
    )

    if not match:
        return ""

    amount = match.group(1)

    plus_sign = (
        match.group(2)
        or ""
    )

    unit = match.group(3)

    raw_value = (
        f"{amount}{plus_sign} "
        f"{unit} ago"
    )

    # "30+ days ago" is not exact.
    # Keep raw value rather than creating a false exact date.

    if plus_sign:

        return raw_value

    return convert_relative_posted_date(
        f"{amount} {unit} ago"
    )

# ============================================================
# EXTRACT INDEED JOB FROM HTML SECTION
# ============================================================

def extract_indeed_job_from_section(
    job_html,
):
    """
    Extract selected Indeed job from:

    <section id="job-full-details">
        ...
    </section>
    """

    if not job_html:

        return {}

    soup = BeautifulSoup(
        job_html,
        "html.parser",
    )

    # =====================================================
    # JOB TITLE
    # =====================================================

    job_title = ""

    title_tag = soup.select_one(
        '[data-testid="jobsearch-JobInfoHeader-title"]'
    )

    if not title_tag:

        title_tag = soup.select_one(
            ".jobsearch-JobInfoHeader-title"
        )

    if title_tag:

        job_title = (
            title_tag.get_text(
                " ",
                strip=True,
            )
        )

        job_title = re.sub(
            r"\s*[-–—]\s*"
            r"job\s+post\s*$",
            "",
            job_title,
            flags=re.IGNORECASE,
        ).strip()

    # =====================================================
    # COMPANY NAME
    # =====================================================

    company_name = ""

    company_tag = soup.select_one(
        '[data-testid="inlineHeader-companyName"]'
    )

    if company_tag:

        company_name = (
            company_tag.get_text(
                " ",
                strip=True,
            )
        )

    # =====================================================
    # LOCATION
    # =====================================================

    location = ""

    location_tag = soup.select_one(
        '[data-testid="inlineHeader-companyLocation"]'
    )

    if location_tag:

        location = (
            location_tag.get_text(
                " ",
                strip=True,
            )
        )

    # =====================================================
    # JOB DESCRIPTION
    # =====================================================

    job_description = ""

    description_tag = soup.find(
        id="jobDescriptionText"
    )

    if description_tag:

        job_description = (
            description_tag.get_text(
                "\n",
                strip=True,
            )
        )

        job_description = re.sub(
            r"\n{3,}",
            "\n\n",
            job_description,
        ).strip()

    # =====================================================
    # COMPLETE SELECTED-JOB TEXT
    # =====================================================

    section_text = soup.get_text(
        "\n",
        strip=True,
    )

    section_text = (
        section_text
        .replace(
            "\xa0",
            " ",
        )
    )

    section_text = re.sub(
        r"\n{3,}",
        "\n\n",
        section_text,
    ).strip()
    # =====================================================
    # JOB KEY + COUNTRY
    # =====================================================

    job_key = ""
    country_code = ""

    apply_widget = soup.select_one(
        "[data-indeed-apply-jk]"
    )

    if apply_widget:

        job_key = (
            apply_widget.get(
                "data-indeed-apply-jk",
                "",
            )
            or ""
        )

        country_code = (
            apply_widget.get(
                "data-indeed-apply-jobcountry",
                "",
            )
            or ""
        )

    # =====================================================
    # APPLY URL
    # =====================================================

    apply_url = ""

    if job_key:

        apply_url = (
            "https://in.indeed.com/"
            f"viewjob?jk={job_key}"
        )

    # =====================================================
    # REMOTE TYPE
    # =====================================================

    remote_type = ""

    lower_location = (
        location.lower()
    )

    lower_section = (
        section_text.lower()
    )

    if (
        "remote" in lower_location
        or "work location: remote"
        in lower_section
    ):

        remote_type = "Remote"

    elif (
        "hybrid" in lower_location
        or "hybrid work"
        in lower_section
    ):

        remote_type = "Hybrid"

    elif (
        location
        and location.lower()
        != "remote"
    ):

        remote_type = "Onsite"

    # =====================================================
    # CLEAN LOCATION
    # =====================================================

    cleaned_location = re.sub(
        r"\s*[•·]\s*"
        r"(?:Remote|Hybrid|On-site|Onsite)"
        r"\s*$",
        "",
        location,
        flags=re.IGNORECASE,
    ).strip()

    if location.lower() == "remote":

        cleaned_location = "Remote"

    # =====================================================
    # CITY / STATE / COUNTRY
    # =====================================================

    (
        city,
        state,
        country,
    ) = parse_indeed_source_location(
        cleaned_location,
        country_code,
    )

    # =====================================================
    # EMPLOYMENT TYPE
    # =====================================================

    employment_type = (
        extract_indeed_source_employment_type(
            section_text
        )
    )

    # =====================================================
    # SALARY
    # =====================================================

    (
        salary_min,
        salary_max,
        salary_currency,
    ) = extract_indeed_source_salary(
        section_text
    )

    # =====================================================
    # EXPERIENCE
    # =====================================================

    (
        experience_required,
        experience_min,
        experience_max,
    ) = extract_experience(
        job_description
    )

    # =====================================================
    # SKILLS
    # =====================================================

    skills = extract_skills(
        job_description
    )

    # =====================================================
    # CATEGORY
    # =====================================================

    category = detect_job_category(
        job_title,
        job_description,
    )

    # =====================================================
    # POSTED DATE
    # =====================================================

    posted_date = (
        extract_indeed_source_posted_date(
            section_text
        )
    )

    return {

        "job_key":
            job_key,

        "job_title":
            job_title,

        "company_name":
            company_name,

        "company_website":
            "",

        "job_description":
            job_description,

        "skills_required":
            skills,

        "experience_required":
            experience_required,

        "experience_min_years":
            experience_min,

        "experience_max_years":
            experience_max,

        "salary_min":
            salary_min,

        "salary_max":
            salary_max,

        "salary_currency":
            salary_currency,

        "location":
            cleaned_location,

        "city":
            city,

        "state":
            state,

        "country":
            country,

        "remote_type":
            remote_type,

        "employment_type":
            employment_type,

        "job_category":
            category,

        "posted_date":
            posted_date,

        "apply_url":
            apply_url,

        "job_details_text":
            section_text,
    }

# ============================================================
# SHINE JOBPOSTING JSON-LD
# ============================================================

def extract_shine_jobposting_json_ld(
    source_html,
):
    """
    Extract Shine Schema.org JobPosting data.

    Example source:

    <script type="application/ld+json">
    {
        "@type": "JobPosting",
        ...
    }
    </script>
    """

    soup = BeautifulSoup(
        source_html,
        "html.parser",
    )

    scripts = soup.find_all(
        "script",
        attrs={
            "type": "application/ld+json",
        },
    )

    for script in scripts:

        raw_json = (
            script.string
            or script.get_text()
        )

        if not raw_json:
            continue

        try:

            data = json.loads(
                raw_json
            )

        except (
            json.JSONDecodeError,
            TypeError,
        ):
            continue

        # -----------------------------------------------
        # Direct JobPosting object
        # -----------------------------------------------

        if (
            isinstance(data, dict)
            and data.get("@type")
            == "JobPosting"
        ):

            return data

        # -----------------------------------------------
        # JobPosting inside a list
        # -----------------------------------------------

        if isinstance(data, list):

            for item in data:

                if (
                    isinstance(item, dict)
                    and item.get("@type")
                    == "JobPosting"
                ):

                    return item

    return {}

# ============================================================
# SHINE NEXT.JS DATA
# ============================================================

def extract_shine_next_data(
    source_html,
):

    soup = BeautifulSoup(
        source_html,
        "html.parser",
    )

    script = soup.find(
        "script",
        id="__NEXT_DATA__",
    )

    if not script:
        return {}

    raw_json = (
        script.string
        or script.get_text()
    )

    if not raw_json:
        return {}

    try:

        return json.loads(
            raw_json
        )

    except (
        json.JSONDecodeError,
        TypeError,
    ):

        return {}

# ============================================================
# FIND SELECTED SHINE JOB
# ============================================================

def find_shine_selected_job(
    next_data,
    expected_job_id="",
):
    """
    Find the selected job object inside __NEXT_DATA__.

    Similar Jobs may also exist in the page data,
    therefore match using the real Shine job ID.
    """

    if not next_data:
        return {}

    expected_job_id = str(
        expected_job_id or ""
    )

    def recursive_find(value):

        if isinstance(
            value,
            dict,
        ):

            # Shine job object usually contains:
            #
            # id
            # jJT
            # jCName

            if (
                "jJT" in value
                and "jCName" in value
            ):

                current_id = str(
                    value.get(
                        "id",
                        "",
                    )
                )

                if (
                    not expected_job_id
                    or current_id
                    == expected_job_id
                ):

                    return value

            for child in value.values():

                result = recursive_find(
                    child
                )

                if result:
                    return result

        elif isinstance(
            value,
            list,
        ):

            for child in value:

                result = recursive_find(
                    child
                )

                if result:
                    return result

        return {}

    return recursive_find(
        next_data
    )

# ============================================================
# SHINE EMPLOYMENT TYPE
# ============================================================

def normalize_shine_employment_type(
    value,
):

    allowed_types = {
        "FULL_TIME": "Full-time",
        "FULL-TIME": "Full-time",
        "FULL TIME": "Full-time",

        "PART_TIME": "Part-time",
        "PART-TIME": "Part-time",
        "PART TIME": "Part-time",

        "INTERNSHIP": "Internship",

        "CONTRACT": "Contract",

        "TEMPORARY": "Temporary",

        "FREELANCE": "Freelance",

        "PERMANENT": "Permanent",
    }

    if not value:
        return ""

    if not isinstance(
        value,
        list,
    ):
        value = [value]

    found_types = []

    for item in value:

        key = str(
            item
        ).strip().upper()

        normalized = (
            allowed_types.get(
                key,
                "",
            )
        )

        if (
            normalized
            and normalized
            not in found_types
        ):

            found_types.append(
                normalized
            )

    return ", ".join(
        found_types
    )

# ============================================================
# SHINE LOCATION
# ============================================================

def extract_shine_source_location(
    jobposting,
):

    cities = []
    states = []
    countries = []

    job_locations = (
        jobposting.get(
            "jobLocation",
            [],
        )
        or []
    )

    if isinstance(
        job_locations,
        dict,
    ):

        job_locations = [
            job_locations
        ]

    country_map = {
        "IN": "India",
        "US": "United States",
        "GB": "United Kingdom",
        "UK": "United Kingdom",
        "CA": "Canada",
        "AU": "Australia",
        "AE": "United Arab Emirates",
        "SG": "Singapore",
    }

    for job_location in job_locations:

        if not isinstance(
            job_location,
            dict,
        ):
            continue

        address = (
            job_location.get(
                "address",
                {},
            )
            or {}
        )

        city = (
            address.get(
                "addressLocality",
                "",
            )
            or ""
        ).strip()

        state = (
            address.get(
                "addressRegion",
                "",
            )
            or ""
        ).strip()

        country = (
            address.get(
                "addressCountry",
                "",
            )
            or ""
        ).strip()

        if city and city not in cities:

            cities.append(
                city
            )

        if state and state not in states:

            states.append(
                state
            )

        if country:

            country = country_map.get(
                country.upper(),
                country,
            )

            if country not in countries:

                countries.append(
                    country
                )

    location = ", ".join(
        cities
    )

    city = ", ".join(
        cities
    )

    state = ", ".join(
        states
    )

    country = ", ".join(
        countries
    )

    return (
        location,
        city,
        state,
        country,
    )

# ============================================================
# ISO DATE NORMALIZER
# ============================================================

def normalize_source_date(
    value,
):

    if not value:
        return ""

    value = str(
        value
    ).strip()

    try:

        parsed = datetime.fromisoformat(
            value.replace(
                "Z",
                "+00:00",
            )
        )

        return parsed.strftime(
            "%Y-%m-%d"
        )

    except ValueError:

        # Safe fallback when value already begins
        # with YYYY-MM-DD.

        match = re.match(
            r"\d{4}-\d{2}-\d{2}",
            value,
        )

        if match:
            return match.group(0)

    return ""

# ============================================================
# SHINE VIEW-SOURCE EXTRACTOR
# ============================================================

def extract_shine_source_job(
    source_html,
):

    # =====================================================
    # VISIBLE SELECTED-JOB TEXT
    # =====================================================

    shine_job_text = (
        isolate_shine_job_details_text(
            source_html
        )
    )

    if not shine_job_text:

        return {}

    # =====================================================
    # 1. JOBPOSTING JSON-LD
    # =====================================================

    jobposting = (
        extract_jobposting_json_ld(
            source_html
        )
    )

    if not jobposting:

        return {}

    # =====================================================
    # 2. JOB ID
    # =====================================================

    identifier = (
        jobposting.get(
            "identifier",
            {},
        )
        or {}
    )

    job_key = str(
        identifier.get(
            "value",
            "",
        )
        or ""
    )

    # =====================================================
    # 3. NEXT.JS DATA
    # =====================================================

    next_data = (
        extract_shine_next_data(
            source_html
        )
    )

    shine_job = (
        find_shine_selected_job(
            next_data,
            job_key,
        )
    )

    # =====================================================
    # 4. JOB TITLE
    # =====================================================

    job_title = (
        jobposting.get(
            "title",
            "",
        )
        or shine_job.get(
            "jJT",
            "",
        )
        or ""
    )

    # =====================================================
    # 5. COMPANY NAME / WEBSITE
    # =====================================================

    organization = (
        jobposting.get(
            "hiringOrganization",
            {},
        )
        or {}
    )

    company_name = (
        organization.get(
            "name",
            "",
        )
        or shine_job.get(
            "jCName",
            "",
        )
        or ""
    )

    company_website = (
        organization.get(
            "sameAs",
            "",
        )
        or organization.get(
            "url",
            "",
        )
        or ""
    )

    # =====================================================
    # 6. DESCRIPTION
    # =====================================================
    # =====================================================
    # JOB DESCRIPTION
    #
    # PRIMARY:
    # visible selected-job section converted to text
    #
    # FALLBACK:
    # JSON-LD description
    # =====================================================

    job_description = (
        extract_shine_job_description_from_text(
            shine_job_text
        )
    )

    # =====================================================
    # FALLBACK TO STRUCTURED DESCRIPTION
    # =====================================================

    if not job_description:

        description_html = (
            jobposting.get(
                "description",
                "",
            )
            or shine_job.get(
                "jJD",
                "",
            )
            or ""
        )

        if description_html:

            description_soup = (
                BeautifulSoup(
                    description_html,
                    "html.parser",
                )
            )

            job_description = (
                description_soup.get_text(
                    "\n",
                    strip=True,
                )
            )

    job_description = re.sub(
        r"\n{3,}",
        "\n\n",
        job_description,
    ).strip()    



    # =====================================================
    # 7. SKILLS
    # =====================================================

    raw_skills = (
        jobposting.get(
            "skills",
            [],
        )
        or []
    )

    if isinstance(
        raw_skills,
        str,
    ):

        raw_skills = re.split(
            r"[,;]+",
            raw_skills,
        )

    skills_list = []

    for skill in raw_skills:

        skill = str(
            skill
        ).strip()

        if (
            skill
            and skill.lower()
            not in {
                item.lower()
                for item in skills_list
            }
        ):

            skills_list.append(
                skill
            )

    if skills_list:

        skills = "; ".join(
            skills_list
        )

    else:

        # Fallback to technology extraction.
        skills = extract_skills(
            job_description
        )

    # =====================================================
    # 8. EXPERIENCE
    # =====================================================

    experience_source = (
        shine_job.get(
            "jExp",
            "",
        )
        or job_description
    )

    (
        experience_required,
        experience_min,
        experience_max,
    ) = extract_experience(
        experience_source
    )

    # =====================================================
    # 9. SALARY
    # =====================================================

    salary_min = ""
    salary_max = ""
    salary_currency = ""

    base_salary = (
        jobposting.get(
            "baseSalary",
            {},
        )
        or {}
    )

    salary_currency = (
        base_salary.get(
            "currency",
            "",
        )
        or ""
    )

    salary_value = (
        base_salary.get(
            "value",
            {},
        )
        or {}
    )

    salary_min = (
        salary_value.get(
            "minValue",
            "",
        )
    )

    salary_max = (
        salary_value.get(
            "maxValue",
            "",
        )
    )

    salary_unit = (
        salary_value.get(
            "unitText",
            "",
        )
        or ""
    ).upper()

    # Monthly -> annual.
    if salary_unit in {
        "MONTH",
        "MONTHLY",
    }:

        if salary_min not in {
            "",
            None,
        }:

            salary_min = (
                float(salary_min)
                * 12
            )

        if salary_max not in {
            "",
            None,
        }:

            salary_max = (
                float(salary_max)
                * 12
            )

    try:

        salary_min = int(
            float(salary_min)
        )

    except (
        TypeError,
        ValueError,
    ):

        salary_min = ""

    try:

        salary_max = int(
            float(salary_max)
        )

    except (
        TypeError,
        ValueError,
    ):

        salary_max = ""

    # =====================================================
    # 10. LOCATION
    # =====================================================

    (
        location,
        city,
        state,
        country,
    ) = extract_shine_source_location(
        jobposting
    )

    # =====================================================
    # 11. REMOTE TYPE
    # =====================================================

    remote_type = ""

    job_location_type = (
        str(
            jobposting.get(
                "jobLocationType",
                "",
            )
            or ""
        )
        .upper()
    )

    work_mode_text = (
        f"{job_title}\n"
        f"{job_description}"
    ).lower()

    if (
        job_location_type
        == "TELECOMMUTE"
        or "work from home"
        in work_mode_text
        or "work from anywhere"
        in work_mode_text
        or re.search(
            r"\bwfh\b",
            work_mode_text,
        )
        or re.search(
            r"\bremote\b",
            work_mode_text,
        )
    ):

        remote_type = "Remote"

    elif "hybrid" in work_mode_text:

        remote_type = "Hybrid"

    elif location:

        remote_type = "Onsite"

    # =====================================================
    # 12. EMPLOYMENT TYPE
    # =====================================================

    employment_type = (
        normalize_shine_employment_type(
            jobposting.get(
                "employmentType"
            )
        )
    )

    # =====================================================
    # 13. JOB CATEGORY
    # =====================================================

    category = detect_job_category(
        job_title,
        job_description,
    )

    # =====================================================
    # 14. POSTED DATE
    # =====================================================

    posted_date = normalize_source_date(
        jobposting.get(
            "datePosted",
            "",
        )
    )

    # =====================================================
    # 15. EXPIRY DATE
    # =====================================================

    expiry_date = normalize_source_date(
        jobposting.get(
            "validThrough",
            "",
        )
    )

    # =====================================================
    # 16. SOURCE URL
    # =====================================================

    extracted_source_url = (
        jobposting.get(
            "url",
            "",
        )
        or ""
    )

    # =====================================================
    # 17. APPLY URL
    # =====================================================

    apply_url = ""

    external_apply_url = (
        shine_job.get(
            "jRUrl",
            "",
        )
        or ""
    )

    if external_apply_url:

        apply_url = external_apply_url

    elif (
        jobposting.get(
            "directApply"
        )
        and extracted_source_url
    ):

        apply_url = (
            extracted_source_url
        )

    return {

        "job_key":
            job_key,

        "job_title":
            job_title,

        "company_name":
            company_name,

        "company_website":
            company_website,

        "job_description":
            job_description,

        "skills_required":
            skills,

        "experience_required":
            experience_required,

        "experience_min_years":
            experience_min,

        "experience_max_years":
            experience_max,

        "salary_min":
            salary_min,

        "salary_max":
            salary_max,

        "salary_currency":
            salary_currency,

        "location":
            location,

        "city":
            city,

        "state":
            state,

        "country":
            country,

        "remote_type":
            remote_type,

        "employment_type":
            employment_type,

        "job_category":
            category,

        "posted_date":
            posted_date,

        "expiry_date":
            expiry_date,

        "application_deadline":
            "",

        "apply_url":
            apply_url,

        "source_url":
            extracted_source_url,

        "job_details_text":
            shine_job_text,
    }

# ============================================================
# HTML -> CLEAN TEXT
# ============================================================

def html_to_clean_text(
    html_content,
):

    if not html_content:
        return ""

    soup = BeautifulSoup(
        html_content,
        "html.parser",
    )

    for unwanted in soup.find_all(
        [
            "script",
            "style",
            "noscript",
            "svg",
        ]
    ):

        unwanted.decompose()

    text = soup.get_text(
        "\n",
        strip=True,
    )

    text = text.replace(
        "\xa0",
        " ",
    )

    text = re.sub(
        r"\n{3,}",
        "\n\n",
        text,
    )

    return text.strip()

# ============================================================
# CUTSHORT SELECTED JOB DETAILS
# ============================================================

def isolate_cutshort_job_details_text(
    source_html,
):
    """
    Cutshort selected-job boundary.

    START:

    <div class="sc-8f06d440-0 jdlUxX">

    END:

    First occurrence AFTER the start of:

    <div class="
        sc-89b45c2f-0
        sc-89b45c2f-1
        cCGhbz
        fERRrh
    ">

    Returns clean plain text, not HTML.
    """

    if not source_html:
        return ""

    # =====================================================
    # START TAG
    # =====================================================

    start_pattern = re.compile(
        r"""
        <div\b
        [^>]*
        class=["']
        [^"']*
        \bsc-8f06d440-0\b
        [^"']*
        \bjdlUxX\b
        [^"']*
        ["']
        [^>]*
        >
        """,
        flags=(
            re.IGNORECASE
            | re.VERBOSE
        ),
    )

    start_match = (
        start_pattern.search(
            source_html
        )
    )

    if not start_match:
        return ""

    # =====================================================
    # END TAG
    #
    # IMPORTANT:
    # search starts AFTER the selected start tag.
    #
    # Therefore this is the FIRST matching end tag
    # belonging to the main selected-job area.
    # =====================================================

    end_pattern = re.compile(
        r"""
        <div\b
        [^>]*
        class=["']
        [^"']*
        \bsc-89b45c2f-0\b
        [^"']*
        \bsc-89b45c2f-1\b
        [^"']*
        \bcCGhbz\b
        [^"']*
        \bfERRrh\b
        [^"']*
        ["']
        [^>]*
        >
        """,
        flags=(
            re.IGNORECASE
            | re.VERBOSE
        ),
    )

    end_match = end_pattern.search(
        source_html,
        start_match.end(),
    )

    if end_match:

        selected_html = source_html[
            start_match.start():
            end_match.start()
        ]

    else:

        # Fallback:
        # if Cutshort changes/removes the end marker,
        # continue from start instead of crashing.

        selected_html = source_html[
            start_match.start():
        ]

    # =====================================================
    # HTML -> CLEAN TEXT
    # =====================================================

    return html_to_clean_text(
        selected_html
    )

# ============================================================
# CUTSHORT NEXT.JS DATA
# ============================================================

def extract_cutshort_next_data(
    source_html,
):

    if not source_html:
        return {}

    soup = BeautifulSoup(
        source_html,
        "html.parser",
    )

    script = soup.find(
        "script",
        id="__NEXT_DATA__",
    )

    if not script:
        return {}

    raw_json = (
        script.string
        or script.get_text()
    )

    if not raw_json:
        return {}

    try:

        return json.loads(
            raw_json
        )

    except (
        json.JSONDecodeError,
        TypeError,
    ):

        return {}

# ============================================================
# CUTSHORT SELECTED PAGE DATA
# ============================================================

def extract_cutshort_page_data(
    source_html,
):

    next_data = (
        extract_cutshort_next_data(
            source_html
        )
    )

    if not next_data:
        return {}

    queries = (
        next_data
        .get(
            "props",
            {},
        )
        .get(
            "pageProps",
            {},
        )
        .get(
            "dehydratedState",
            {},
        )
        .get(
            "queries",
            [],
        )
        or []
    )

    # Search every query instead of assuming
    # the selected job is always query number 0.

    for query in queries:

        data = (
            query
            .get(
                "state",
                {},
            )
            .get(
                "data",
                {},
            )
            .get(
                "data",
                {},
            )
            or {}
        )

        page_data = (
            data.get(
                "pageData"
            )
        )

        if (
            isinstance(
                page_data,
                dict,
            )
            and (
                page_data.get(
                    "_id"
                )
                or page_data.get(
                    "headline"
                )
            )
        ):

            return page_data

    return {}

# ============================================================
# CUTSHORT EMPLOYMENT TYPE
# ============================================================

def normalize_cutshort_employment_type(
    values,
):

    allowed_types = {
        "full_time": "Full-time",
        "full-time": "Full-time",
        "full time": "Full-time",

        "part_time": "Part-time",
        "part-time": "Part-time",
        "part time": "Part-time",

        "internship": "Internship",

        "contract": "Contract",

        "temporary": "Temporary",

        "freelance": "Freelance",

        "permanent": "Permanent",
    }

    if not values:
        return ""

    if not isinstance(
        values,
        list,
    ):
        values = [values]

    found_types = []

    for value in values:

        key = (
            str(value)
            .strip()
            .lower()
        )

        employment_type = (
            allowed_types.get(
                key,
                "",
            )
        )

        if (
            employment_type
            and employment_type
            not in found_types
        ):

            found_types.append(
                employment_type
            )

    return ", ".join(
        found_types
    )

# ============================================================
# CUTSHORT CITY / STATE / COUNTRY
# ============================================================

def parse_cutshort_location(
    location,
    job_description,
    remote_type,
):

    city = ""
    state = ""
    country = ""

    # =====================================================
    # LOOK FOR DESCRIPTION LOCATION
    #
    # Example:
    #
    # Location
    # : Pune Remote
    # =====================================================

    location_match = re.search(
        r"""
        \bLocation
        \s*
        :?
        \s*
        ([^\n]+)
        """,
        job_description,
        flags=(
            re.IGNORECASE
            | re.VERBOSE
        ),
    )

    if location_match:

        description_location = (
            location_match.group(1)
            .strip()
        )

        description_location = re.sub(
            r"\b"
            r"(Remote|Hybrid|Onsite|On-site)"
            r"\b",
            "",
            description_location,
            flags=re.IGNORECASE,
        ).strip(
            " ,-"
        )

        if description_location:

            parts = [
                part.strip()
                for part
                in description_location.split(",")
                if part.strip()
            ]

            if len(parts) >= 2:

                city = ", ".join(
                    parts[:-1]
                )

                state = parts[-1]

            elif len(parts) == 1:

                city = parts[0]

    # =====================================================
    # FALLBACK TO MAIN CUTSHORT LOCATION
    # =====================================================

    if (
        not city
        and location
        and remote_type != "Remote"
    ):

        parts = [
            part.strip()
            for part
            in location.split(",")
            if part.strip()
        ]

        if len(parts) >= 2:

            city = ", ".join(
                parts[:-1]
            )

            state = parts[-1]

        elif len(parts) == 1:

            city = parts[0]

    return (
        city,
        state,
        country,
    )

# ============================================================
# CUTSHORT VIEW-SOURCE EXTRACTOR
# ============================================================

def extract_cutshort_source_job(
    source_html,
):

    # =====================================================
    # 1. SELECTED JOB TEXT USING YOUR EXACT BOUNDARY
    # =====================================================

    job_details_text = (
        isolate_cutshort_job_details_text(
            source_html
        )
    )

    # =====================================================
    # 2. CUTSHORT INTERNAL PAGE DATA
    # =====================================================

    page_data = (
        extract_cutshort_page_data(
            source_html
        )
    )

    # =====================================================
    # 3. GENERIC SCHEMA.ORG JOBPOSTING
    #
    # Reuse existing common function.
    # =====================================================

    jobposting = (
        extract_jobposting_json_ld(
            source_html
        )
    )

    if (
        not job_details_text
        and not page_data
        and not jobposting
    ):

        return {}

    # =====================================================
    # JOB KEY
    # =====================================================

    identifier = (
        jobposting.get(
            "identifier",
            {},
        )
        or {}
    )

    job_key = str(
        page_data.get(
            "_id"
        )
        or identifier.get(
            "value"
        )
        or ""
    )

    # =====================================================
    # JOB TITLE
    # =====================================================

    job_title = (
        page_data.get(
            "headline"
        )
        or jobposting.get(
            "title"
        )
        or ""
    )

    # JSON-LD can contain:
    #
    # MERN Stack Engineer (SDE-2) (Remote)
    #
    # Remote already has its own Excel column.

    job_title = re.sub(
        r"\s*\(Remote\)\s*$",
        "",
        job_title,
        flags=re.IGNORECASE,
    ).strip()

    # =====================================================
    # COMPANY NAME
    # =====================================================

    company_id = (
        page_data.get(
            "companyId",
            {},
        )
        or {}
    )

    organization = (
        jobposting.get(
            "hiringOrganization",
            {},
        )
        or {}
    )

    company_name = (
        company_id.get(
            "name"
        )
        or organization.get(
            "name"
        )
        or ""
    )

    # =====================================================
    # COMPANY WEBSITE
    # =====================================================

    company_website = ""

    company_details = (
        page_data.get(
            "companyDetails",
            {},
        )
        or {}
    )

    links = (
        company_details.get(
            "links"
        )
        or (
            company_details.get(
                "company",
                {},
            )
            or {}
        ).get(
            "links"
        )
        or {}
    )

    if isinstance(
        links,
        dict,
    ):

        company_website = (
            links.get(
                "website",
                "",
            )
            or ""
        )

    # =====================================================
    # JOB DESCRIPTION
    # =====================================================

    description_html = (
        page_data.get(
            "sanitizedComment"
        )
        or jobposting.get(
            "description"
        )
        or ""
    )

    job_description = (
        html_to_clean_text(
            unescape(
                description_html
            )
        )
    )

    # =====================================================
    # SKILLS
    # =====================================================

    raw_skills = (
        page_data.get(
            "allSkills"
        )
        or jobposting.get(
            "skills"
        )
        or []
    )

    if isinstance(
        raw_skills,
        str,
    ):

        raw_skills = re.split(
            r"[,;]+",
            raw_skills,
        )

    skills_list = []

    for skill in raw_skills:

        skill = str(
            skill
        ).strip()

        if (
            skill
            and skill not in skills_list
        ):

            skills_list.append(
                skill
            )

    skills = "; ".join(
        skills_list
    )

    if not skills:

        skills = extract_skills(
            job_description
        )

    # =====================================================
    # EXPERIENCE
    # =====================================================

    exp_range = (
        page_data.get(
            "expRange",
            {},
        )
        or {}
    )

    experience_min = (
        exp_range.get(
            "minVanity"
        )
    )

    if experience_min in {
        "",
        None,
    }:

        experience_min = (
            exp_range.get(
                "min",
                "",
            )
        )

    experience_max = (
        exp_range.get(
            "maxVanity"
        )
    )

    if experience_max in {
        "",
        None,
    }:

        experience_max = (
            exp_range.get(
                "max",
                "",
            )
        )

    experience_required = ""

    if (
        experience_min not in {
            "",
            None,
        }
        and experience_max not in {
            "",
            None,
        }
    ):

        experience_required = (
            f"{experience_min} "
            f"to {experience_max} Years"
        )

    # Fallback to selected job text.
    if not experience_required:

        (
            experience_required,
            experience_min,
            experience_max,
        ) = extract_experience(
            job_details_text
        )

    # =====================================================
    # SALARY
    # =====================================================

    salary_min = ""
    salary_max = ""
    salary_currency = ""

    base_salary = (
        jobposting.get(
            "baseSalary",
            {},
        )
        or {}
    )

    salary_currency = (
        base_salary.get(
            "currency",
            "",
        )
        or ""
    )

    salary_value = (
        base_salary.get(
            "value",
            {},
        )
        or {}
    )

    salary_min = (
        salary_value.get(
            "minValue",
            "",
        )
    )

    salary_max = (
        salary_value.get(
            "maxValue",
            "",
        )
    )

    salary_unit = (
        salary_value.get(
            "unitText",
            "",
        )
        or ""
    ).upper()

    # Monthly -> annual normalization.

    if salary_unit in {
        "MONTH",
        "MONTHLY",
    }:

        if salary_min not in {
            "",
            None,
        }:

            salary_min = (
                float(salary_min)
                * 12
            )

        if salary_max not in {
            "",
            None,
        }:

            salary_max = (
                float(salary_max)
                * 12
            )

    try:

        salary_min = int(
            float(
                salary_min
            )
        )

    except (
        TypeError,
        ValueError,
    ):

        salary_min = ""

    try:

        salary_max = int(
            float(
                salary_max
            )
        )

    except (
        TypeError,
        ValueError,
    ):

        salary_max = ""

    # =====================================================
    # LOCATION / REMOTE TYPE
    # =====================================================

    location = str(
        page_data.get(
            "locations",
            "",
        )
        or ""
    ).strip()

    remote_value = str(
        page_data.get(
            "remoteType",
            "",
        )
        or ""
    ).lower()

    job_location_type = str(
        jobposting.get(
            "jobLocationType",
            "",
        )
        or ""
    ).upper()

    remote_type = ""

    if (
        "remote" in remote_value
        or job_location_type
        == "TELECOMMUTE"
        or "remote only"
        in location.lower()
    ):

        remote_type = "Remote"

    elif "hybrid" in remote_value:

        remote_type = "Hybrid"

    elif location:

        remote_type = "Onsite"

    (
        city,
        state,
        country,
    ) = parse_cutshort_location(
        location,
        job_description,
        remote_type,
    )

    # =====================================================
    # EMPLOYMENT TYPE
    # =====================================================

    employment_type = (
        normalize_cutshort_employment_type(
            page_data.get(
                "roleTypes"
            )
            or jobposting.get(
                "employmentType"
            )
        )
    )

    # =====================================================
    # CATEGORY
    # =====================================================

    category = detect_job_category(
        job_title,
        job_description,
    )

    # Cutshort itself categorizes this as tech.
    if not category:

        tag_category = (
            page_data
            .get(
                "matchPreferences",
                {},
            )
            .get(
                "tagCategory",
                "",
            )
        )

        if (
            str(tag_category)
            .lower()
            == "tech"
        ):

            category = "IT"

    # =====================================================
    # POSTED / EXPIRY
    # =====================================================

    posted_date = (
        normalize_source_date(
            jobposting.get(
                "datePosted",
                "",
            )
        )
    )

    expiry_date = (
        normalize_source_date(
            jobposting.get(
                "validThrough",
                "",
            )
        )
    )

    # =====================================================
    # SOURCE URL
    # =====================================================

    extracted_source_url = (
        page_data.get(
            "publicUrl",
            "",
        )
        or ""
    )

    if not extracted_source_url:

        soup = BeautifulSoup(
            source_html,
            "html.parser",
        )

        canonical = soup.find(
            "link",
            rel="canonical",
        )

        if canonical:

            extracted_source_url = (
                canonical.get(
                    "href",
                    "",
                )
                or ""
            )

    # =====================================================
    # APPLY URL
    # =====================================================

    apply_url = ""

    if (
        jobposting.get(
            "directApply"
        )
        and extracted_source_url
    ):

        apply_url = (
            extracted_source_url
        )

    # =====================================================
    # RETURN COMMON NORMALIZED FORMAT
    # =====================================================

    return {

        "job_key":
            job_key,

        "job_title":
            job_title,

        "company_name":
            company_name,

        "company_website":
            company_website,

        "job_description":
            job_description,

        "skills_required":
            skills,

        "experience_required":
            experience_required,

        "experience_min_years":
            experience_min,

        "experience_max_years":
            experience_max,

        "salary_min":
            salary_min,

        "salary_max":
            salary_max,

        "salary_currency":
            salary_currency,

        "location":
            location,

        "city":
            city,

        "state":
            state,

        "country":
            country,

        "remote_type":
            remote_type,

        "employment_type":
            employment_type,

        "job_category":
            category,

        "posted_date":
            posted_date,

        "expiry_date":
            expiry_date,

        "application_deadline":
            "",

        "apply_url":
            apply_url,

        "source_url":
            extracted_source_url,

        "job_details_text":
            job_details_text,
    }


# ============================================================
# PORTAL PARSER DISPATCHER
# ============================================================

def extract_portal_job(
    source_platform,
    source_html,
    normalized_text,
):

    # =====================================================
    # INDEED
    # =====================================================

    if source_platform == "Indeed":

        indeed_job_html = (
            isolate_indeed_job_section(
                source_html
            )
        )

        if not indeed_job_html:

            raise ValueError(
                "Indeed selected-job section "
                "could not be found."
            )

        return (
            extract_indeed_job_from_section(
                indeed_job_html
            )
        )

    # =====================================================
    # SHINE
    # =====================================================

    if source_platform == "Shine":

        return (
            extract_shine_source_job(
                source_html
            )
        )
    # =====================================================
    # CUTSHORT
    # =====================================================
    if source_platform == "Cutshort":
        cutshort_data = (
            extract_cutshort_source_job(
                source_html
            )
        )
        if not cutshort_data:
            raise ValueError(
                "Cutshort selected job data "
                "could not be extracted."
            )

        return cutshort_data

    # =====================================================
    # LINKEDIN
    #
    # Currently old text parser.
    # Later replace with View-Source parser.
    # =====================================================

    if source_platform == "LinkedIn":

        return (
            extract_linkedin_normalized_job(
                normalized_text
            )
        )

    # =====================================================
    # FUTURE
    # =====================================================

    # if source_platform == "Naukri":
    #     return extract_naukri_source_job(
    #         source_html
    #     )

    # if source_platform == "Foundit":
    #     return extract_foundit_source_job(
    #         source_html
    #     )

    return {}

def extract_linkedin_normalized_job(
    text,
):

    result = empty_extracted_job()

    main_text = (
        isolate_linkedin_job_text(
            text
        )
    )

    (
        job_title,
        company_name,
        location,
        posted_date_raw,
    ) = extract_linkedin_header(
        main_text
    )

    job_description = (
        extract_linkedin_job_description(
            main_text
        )
    )

    (
        experience_required,
        experience_min,
        experience_max,
    ) = extract_experience(
        job_description
    )

    if not experience_required:

        (
            experience_required,
            experience_min,
            experience_max,
        ) = extract_experience(
            main_text
        )

    (
        salary_min,
        salary_max,
        salary_currency,
    ) = extract_salary(
        main_text
    )

    (
        city,
        state,
        country,
    ) = parse_linkedin_location(
        location
    )

    skills = extract_skills(
        job_description
    )

    remote_type = (
        detect_linkedin_remote_type(
            main_text
        )
    )

    employment_type = (
        extract_linkedin_employment_type(
            main_text
        )
    )

    category = (
        extract_linkedin_job_category(
            main_text,
            job_title,
            job_description,
        )
    )

    posted_date = (
        convert_relative_posted_date(
            posted_date_raw
        )
    )

    company_website = (
        extract_company_website(
            main_text
        )
    )

    apply_url = get_label_value(
        main_text,
        [
            "Apply URL",
            "Application URL",
        ],
    )

    result.update({

        "job_title":
            job_title,

        "company_name":
            company_name,

        "company_website":
            company_website,

        "job_description":
            job_description,

        "skills_required":
            skills,

        "experience_required":
            experience_required,

        "experience_min_years":
            experience_min,

        "experience_max_years":
            experience_max,

        "salary_min":
            salary_min,

        "salary_max":
            salary_max,

        "salary_currency":
            salary_currency,

        "location":
            location,

        "city":
            city,

        "state":
            state,

        "country":
            country,

        "remote_type":
            remote_type,

        "employment_type":
            employment_type,

        "job_category":
            category,

        "posted_date":
            posted_date,

        "apply_url":
            apply_url,

        "job_details_text":
            main_text,
    })

    return result


# ============================================================
# ISOLATE SHINE SELECTED JOB AS PLAIN TEXT
# ============================================================

def isolate_shine_job_details_text(
    source_html,
):
    """
    Extract only the selected Shine job left-side content.

    Start area looks like:

    <div class="jdLeft_jdBodyLeft__KdoJ4">
        <div class="jdLeft_jdLeft__GZjqp">
            ...
        </div>
    </div>

    The right side:

    <aside class="jdRight_jdBodyRight__qTz6e">

    contains Similar Jobs and must NOT be included.

    Returns PLAIN TEXT, not HTML.
    """

    if not source_html:
        return ""

    soup = BeautifulSoup(
        source_html,
        "html.parser",
    )

    # =====================================================
    # FIND OUTER LEFT JOB CONTAINER
    # =====================================================
    #
    # Do not use the complete class:
    #
    # jdLeft_jdBodyLeft__KdoJ4
    #
    # because KdoJ4 may change.
    #
    # Match only stable prefix:
    #
    # jdLeft_jdBodyLeft__
    # =====================================================

    outer_left = soup.find(
        lambda tag:
            tag.name == "div"
            and any(
                class_name.startswith(
                    "jdLeft_jdBodyLeft__"
                )
                for class_name
                in tag.get(
                    "class",
                    [],
                )
            )
    )

    if not outer_left:
        return ""

    # =====================================================
    # FIND INNER LEFT JOB DETAILS
    # =====================================================

    inner_left = outer_left.find(
        lambda tag:
            tag.name == "div"
            and any(
                class_name.startswith(
                    "jdLeft_jdLeft__"
                )
                for class_name
                in tag.get(
                    "class",
                    [],
                )
            )
    )

    # Prefer inner container when available.
    job_container = (
        inner_left
        or outer_left
    )

    # =====================================================
    # REMOVE NON-VISIBLE / USELESS ELEMENTS
    # =====================================================

    for unwanted in job_container.find_all(
        [
            "script",
            "style",
            "noscript",
            "svg",
        ]
    ):

        unwanted.decompose()

    # =====================================================
    # HTML -> PLAIN TEXT
    # =====================================================

    job_text = (
        job_container.get_text(
            "\n",
            strip=True,
        )
    )

    job_text = (
        job_text
        .replace(
            "\xa0",
            " ",
        )
    )

    # Remove excessive empty lines.
    job_text = re.sub(
        r"\n{3,}",
        "\n\n",
        job_text,
    )

    return job_text.strip()


# ============================================================
# SHINE JOB DESCRIPTION FROM LEFT-SIDE TEXT
# ============================================================

def extract_shine_job_description_from_text(
    shine_job_text,
):
    """
    Extract plain-text description from the Shine
    selected-job left panel.

    Start:
        JOB DESCRIPTION

    Stop before:
        Other Details
        Recruiter Details
        Company Details
        About Recruiter
        About Company
    """

    if not shine_job_text:
        return ""

    description = extract_section(
        shine_job_text,
        "JOB DESCRIPTION",
        [
            "Other Details",
            "Recruiter Details",
            "Company Details",
            "About Recruiter",
            "About Company",
        ],
    )

    description = re.sub(
        r"\n{3,}",
        "\n\n",
        description,
    )

    return description.strip()
# ============================================================
# MAIN EXTRACTION
# ============================================================

def extract_job_data(
    raw_text,
    source_url,
):

    # =====================================================
    # 1. RECONSTRUCT ORIGINAL SOURCE
    # =====================================================

    source_html = (
        unwrap_view_source_file(
            raw_text
        )
    )

    # =====================================================
    # 2. TEXT VERSION FOR FALLBACK PARSERS
    # =====================================================

    normalized_text = (
        normalize_text(
            source_html
        )
    )

    # =====================================================
    # 3. DETECT PLATFORM
    # =====================================================

    source_platform = (
        detect_source_platform(
            source_html,
            source_url,
        )
    )

    # Fallback to original TXT if needed.
    if source_platform == "Other":

        source_platform = (
            detect_source_platform(
                raw_text,
                source_url,
            )
        )

    # =====================================================
    # 4. CALL CORRECT PORTAL ADAPTER
    # =====================================================

    extracted = (
        extract_portal_job(
            source_platform,
            source_html,
            normalized_text,
        )
    )

    # =====================================================
    # 5. GENERIC FALLBACK
    # =====================================================

    if not extracted:

        extracted = (
            extract_generic_job(
                source_html,
                normalized_text,
            )
        )

    # =====================================================
    # 6. NORMALIZED VALUES
    # =====================================================

    job_title = extracted.get(
        "job_title",
        "",
    )

    company_name = extracted.get(
        "company_name",
        "",
    )

    company_website = extracted.get(
        "company_website",
        "",
    )

    job_description = extracted.get(
        "job_description",
        "",
    )

    skills = extracted.get(
        "skills_required",
        "",
    )

    experience_required = extracted.get(
        "experience_required",
        "",
    )

    experience_min = extracted.get(
        "experience_min_years",
        "",
    )

    experience_max = extracted.get(
        "experience_max_years",
        "",
    )

    salary_min = extracted.get(
        "salary_min",
        "",
    )

    salary_max = extracted.get(
        "salary_max",
        "",
    )

    salary_currency = extracted.get(
        "salary_currency",
        "",
    )

    location = extracted.get(
        "location",
        "",
    )

    city = extracted.get(
        "city",
        "",
    )

    state = extracted.get(
        "state",
        "",
    )

    country = extracted.get(
        "country",
        "",
    )

    country = infer_country_from_location(
        location=location,
        state=state,
        existing_country=country,
    )

    remote_type = extracted.get(
        "remote_type",
        "",
    )

    employment_type = extracted.get(
        "employment_type",
        "",
    )

    category = extracted.get(
        "job_category",
        "",
    )

    posted_date = extracted.get(
        "posted_date",
        "",
    )

    expiry_date = extracted.get(
        "expiry_date",
        "",
    )

    application_deadline = (
        extracted.get(
            "application_deadline",
            "",
        )
    )

    apply_url = extracted.get(
        "apply_url",
        "",
    )

    # =====================================================
    # SOURCE URL
    # =====================================================

    if not source_url:

        source_url = extracted.get(
            "source_url",
            "",
        )

    # =====================================================
    # MAIN TEXT FOR FALLBACK HASH
    # =====================================================

    main_text = (
        extracted.get(
            "job_details_text",
            "",
        )
        or job_description
    )

    # =====================================================
    # SOURCE TYPE / SCORE
    # =====================================================

    source_type = (
        determine_source_type(
            source_platform
        )
    )

    score = (
        get_source_score(
            source_type
        )
    )

    # =====================================================
    # STABLE PLATFORM JOB ID
    # =====================================================

    platform_job_key = (
        extracted.get(
            "job_key",
            "",
        )
    )

    if platform_job_key:

        internal_job_id = str(
            platform_job_key
        )

    else:

        internal_job_id = (
            generate_internal_job_id(
                main_text,
                source_url,
            )
        )

    # =====================================================
    # COMMON JOB ID
    # =====================================================

    platform_prefix = (
        source_platform
        .lower()
        .replace(
            " ",
            "_",
        )
    )

    job_id = (
        f"{platform_prefix}_"
        f"{internal_job_id[:16]}"
    )

    now = (
        timezone.now()
        .isoformat()
    )

    # =====================================================
    # FINAL COMMON EXCEL ROW
    # =====================================================

    return {

        "job_id":
            job_id,

        "internal_job_id":
            internal_job_id,

        "job_title":
            job_title,

        "company_name":
            company_name,

        "company_website":
            company_website,

        "job_description":
            job_description,

        "skills_required":
            skills,

        "experience_required":
            experience_required,

        "experience_min_years":
            experience_min,

        "experience_max_years":
            experience_max,

        "salary_min":
            salary_min,

        "salary_max":
            salary_max,

        "salary_currency":
            salary_currency,

        "location":
            location,

        "city":
            city,

        "state":
            state,

        "country":
            country,

        "remote_type":
            remote_type,

        "employment_type":
            employment_type,

        "job_category":
            category,

        "posted_date":
            posted_date,

        "expiry_date":
            expiry_date,

        "application_deadline":
            application_deadline,

        "apply_url":
            apply_url,

        "updated_at":
            now,

        "fetched_at":
            now,

        "is_fake":
            False,

        "is_active":
            True,

        "source_url":
            source_url,

        "source_type":
            source_type,

        "source_platform":
            source_platform,

        "score":
            score,
    }

def extract_generic_job(
    source_html,
    normalized_text,
):

    result = empty_extracted_job()

    jobposting = (
        extract_jobposting_json_ld(
            source_html
        )
    )

    if not jobposting:

        return result

    # =====================================================
    # TITLE
    # =====================================================

    result["job_title"] = (
        jobposting.get(
            "title",
            "",
        )
        or ""
    )

    # =====================================================
    # COMPANY
    # =====================================================

    organization = (
        jobposting.get(
            "hiringOrganization",
            {},
        )
        or {}
    )

    result["company_name"] = (
        organization.get(
            "name",
            "",
        )
        or ""
    )

    result["company_website"] = (
        organization.get(
            "sameAs",
            "",
        )
        or organization.get(
            "url",
            "",
        )
        or ""
    )

    # =====================================================
    # DESCRIPTION
    # =====================================================

    description_html = (
        jobposting.get(
            "description",
            "",
        )
        or ""
    )

    result["job_description"] = (
        html_to_clean_text(
            description_html
        )
    )

    # =====================================================
    # DATES
    # =====================================================

    result["posted_date"] = (
        normalize_source_date(
            jobposting.get(
                "datePosted",
                "",
            )
        )
    )

    result["expiry_date"] = (
        normalize_source_date(
            jobposting.get(
                "validThrough",
                "",
            )
        )
    )

    # =====================================================
    # URL
    # =====================================================

    result["source_url"] = (
        jobposting.get(
            "url",
            "",
        )
        or ""
    )

    # =====================================================
    # CATEGORY
    # =====================================================

    result["job_category"] = (
        detect_job_category(
            result["job_title"],
            result["job_description"],
        )
    )

    # =====================================================
    # SKILLS FALLBACK
    # =====================================================

    result["skills_required"] = (
        extract_skills(
            result[
                "job_description"
            ]
        )
    )

    result["job_details_text"] = (
        result[
            "job_description"
        ]
    )

    return result


# ============================================================
# CREATE EXCEL
# ============================================================

def create_excel(job_data):

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = "Job Data"

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1E3A8A",
    )

    header_font = Font(
        bold=True,
        color="FFFFFF",
    )

    # Header row
    for column_number, column_name in enumerate(
        JOB_COLUMNS,
        start=1,
    ):

        cell = worksheet.cell(
            row=1,
            column=column_number,
            value=column_name,
        )

        cell.fill = header_fill
        cell.font = header_font

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    # Job row
    for column_number, column_name in enumerate(
        JOB_COLUMNS,
        start=1,
    ):

        cell = worksheet.cell(
            row=2,
            column=column_number,
            value=job_data.get(
                column_name,
                "",
            ),
        )

        cell.alignment = Alignment(
            vertical="top",
            wrap_text=True,
        )

    worksheet.freeze_panes = "A2"

    last_column = get_column_letter(
        len(JOB_COLUMNS)
    )

    worksheet.auto_filter.ref = (
        f"A1:{last_column}2"
    )

    large_columns = {
        "job_description",
        "skills_required",
        "location",
        "source_url",
        "apply_url",
    }

    for column_number, column_name in enumerate(
        JOB_COLUMNS,
        start=1,
    ):

        letter = get_column_letter(
            column_number
        )

        if column_name in large_columns:
            width = 45
        else:
            width = 20

        worksheet.column_dimensions[
            letter
        ].width = width

    worksheet.row_dimensions[1].height = 35
    worksheet.row_dimensions[2].height = 120

    output = BytesIO()

    workbook.save(output)

    output.seek(0)

    return output


# ============================================================
# DJANGO VIEW
# ============================================================

def upload_job_text(request):

    if request.method == "POST":

        form = JobTextUploadForm(
            request.POST,
            request.FILES,
        )

        if form.is_valid():

            uploaded_file = (
                form.cleaned_data[
                    "text_file"
                ]
            )

            source_url = (
                form.cleaned_data.get(
                    "source_url",
                    "",
                )
            )

            raw_text = (
                decode_uploaded_file(
                    uploaded_file
                )
            )

            job_data = extract_job_data(
                raw_text,
                source_url,
            )

            excel_file = create_excel(
                job_data
            )

            job_title = (
                job_data.get(
                    "job_title"
                )
                or Path(
                    uploaded_file.name
                ).stem
            )

            safe_name = (
                slugify(job_title)
                or "job"
            )

            filename = (
                f"{safe_name}_"
                f"job_data.xlsx"
            )

            response = HttpResponse(
                excel_file.getvalue(),
                content_type=(
                    "application/vnd."
                    "openxmlformats-officedocument."
                    "spreadsheetml.sheet"
                ),
            )

            response[
                "Content-Disposition"
            ] = (
                f'attachment; '
                f'filename="{filename}"'
            )

            return response

    else:

        form = JobTextUploadForm()

    return render(
        request,
        "extractor/upload.html",
        {
            "form": form,
        },
    )
